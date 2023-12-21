import openpyxl

class HomePageData:

    test_HomePage_Data = [{"fname":"Vipin","email":"Vipin@gmail.com","gend":"Male"},
                          {"fname":"Vanshika","email":"Vanshika@gmail.com","gend":"Female"}]

    @staticmethod
    def getExcelData(sheet_name):
        Dict = {}
        wb = openpyxl.load_workbook("../SampleExcel.xlsx")
        ws = wb.active
        total_row = ws.max_row
        total_col = ws.max_column

        for i in range(1,total_row+1):
            if ws.cell(row=i, column=1).value == sheet_name:
                for j in range(2, total_col+1):
                    Dict[ws.cell(row=1,column=j).value] = ws.cell(row=i,column=j).value
                    print(Dict)
        return [Dict]






